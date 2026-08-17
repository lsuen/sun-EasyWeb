"""
数据驱动层 - 支持Excel和JSON数据格式 (标准字段版)

作者：孙文龙 | 许可证：MIT
"""
import os
import re
import json
from typing import List, Dict, Any, Optional

from utils.logger import get_logger


class DataDriver:
    """数据驱动器 - 读取Excel/JSON测试数据"""
    
    # 标准必填字段
    REQUIRED_FIELDS = ['id', 'name', 'type', 'url', 'expected_type', 'expected_value']
    
    def __init__(self, data_path: str):
        self.data_path = data_path
        self.logger = get_logger('DataDriver')
        self.data: List[Dict[str, Any]] = []
    
    def load(self) -> List[Dict[str, Any]]:
        """加载数据文件"""
        if not os.path.exists(self.data_path):
            raise FileNotFoundError(f"数据文件不存在: {self.data_path}")
        
        self.logger.info(f"加载数据文件: {self.data_path}")
        
        if self.data_path.endswith(('.xlsx', '.xls')):
            self.data = self._load_excel()
        elif self.data_path.endswith('.json'):
            self.data = self._load_json()
        else:
            raise ValueError(f"不支持的数据格式: {self.data_path}")
        
        # 校验字段
        self._validate_fields()
        
        self.logger.info(f"数据加载成功: {len(self.data)} 条记录")
        return self.data
    
    def _load_excel(self) -> List[Dict[str, Any]]:
        """加载Excel文件"""
        import openpyxl
        
        workbook = openpyxl.load_workbook(self.data_path)
        all_data = []
        
        # 遍历所有 Sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 读取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else '')
            
            # 读取数据行
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                
                # 跳过空行
                if any(v is not None for v in row_dict.values()):
                    # 补充缺失字段为空字符串
                    for field in self.REQUIRED_FIELDS:
                        if field not in row_dict:
                            row_dict[field] = ''
                    all_data.append(row_dict)
        
        workbook.close()
        return all_data
    
    def _load_json(self) -> List[Dict[str, Any]]:
        """加载JSON文件"""
        with open(self.data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 支持列表或字典格式
        if isinstance(data, dict):
            data = list(data.values())[0] if data else []
        
        # 补充缺失字段
        for row in data:
            for field in self.REQUIRED_FIELDS:
                if field not in row:
                    row[field] = ''
        
        return data
    
    def _validate_fields(self):
        """校验必填字段"""
        errors = []
        for i, row in enumerate(self.data):
            for field in self.REQUIRED_FIELDS:
                if field not in row or not row[field]:
                    errors.append(f"用例 {row.get('id', i+1)} 缺少必填字段: {field}")
        
        if errors:
            self.logger.warning("字段校验警告:")
            for err in errors:
                self.logger.warning(f"  - {err}")
            # 不抛出异常，允许部分字段缺失，但记录警告
    
    def get(self, index: int) -> Optional[Dict[str, Any]]:
        """获取指定索引的数据"""
        if 0 <= index < len(self.data):
            return self.data[index]
        return None
    
    def __len__(self) -> int:
        return len(self.data)
    
    def __iter__(self):
        return iter(self.data)
    
    @staticmethod
    def replace_placeholders(data: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, Any]:
        """
        替换数据中的占位符 ${var}
        
        示例:
            data = {'url': '${base_url}/login', 'username': 'test'}
            context = {'base_url': 'https://example.com'}
            结果: {'url': 'https://example.com/login', 'username': 'test'}
        """
        result = {}
        pattern = re.compile(r'\$\{(\w+)\}')
        
        for key, value in data.items():
            if isinstance(value, str):
                def replace_match(match):
                    var_name = match.group(1)
                    return str(context.get(var_name, match.group(0)))
                
                result[key] = pattern.sub(replace_match, value)
            else:
                result[key] = value
        
        return result
